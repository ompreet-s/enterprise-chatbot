import os
from dotenv import load_dotenv

load_dotenv()

# ── LAZY LOADING — models load only when first used ───────────────────────────
_llm = None
_embeddings = None
_reranker = None

def get_llm():
    global _llm
    if _llm is None:
        from langchain_groq import ChatGroq
        _llm = ChatGroq(
            model="llama-3.3-70b-versatile",
            temperature=0.1,
            api_key=os.getenv("GROQ_API_KEY")
        )
        print("✅ LLM loaded")
    return _llm

def get_embeddings():
    global _embeddings
    if _embeddings is None:
        from langchain_community.embeddings import HuggingFaceEmbeddings
        _embeddings = HuggingFaceEmbeddings(
            model_name="BAAI/bge-large-en-v1.5",
            encode_kwargs={"normalize_embeddings": True}
        )
        print("✅ Embeddings loaded")
    return _embeddings

def get_reranker():
    global _reranker
    if _reranker is None:
        from sentence_transformers import CrossEncoder
        _reranker = CrossEncoder("cross-encoder/ms-marco-MiniLM-L-6-v2")
        print("✅ Reranker loaded")
    return _reranker

# ── SPLITTER ──────────────────────────────────────────────────────────────────

def get_splitter():
    from langchain_text_splitters import RecursiveCharacterTextSplitter
    return RecursiveCharacterTextSplitter(
        chunk_size=1000,
        chunk_overlap=200,
        separators=["\n\n\n", "\n\n", "\n", ". ", " "],
    )

# ── DOCUMENT LOADER ───────────────────────────────────────────────────────────

def load_document(file_path, filename):
    ext = filename.split(".")[-1].lower()

    if ext == "pdf":
        import fitz
        doc = fitz.open(file_path)
        pages = [
            f"[Page {i+1}]\n{p.get_text()}"
            for i, p in enumerate(doc) if p.get_text().strip()
        ]
        return "\n\n".join(pages)

    elif ext == "docx":
        from docx import Document as DocxDoc
        doc = DocxDoc(file_path)
        return "\n\n".join([p.text for p in doc.paragraphs if p.text.strip()])

    elif ext == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        lines = []
        for sheet in wb.sheetnames:
            lines.append(f"--- Sheet: {sheet} ---")
            for row in wb[sheet].iter_rows(values_only=True):
                r = " | ".join([str(c) for c in row if c is not None])
                if r.strip():
                    lines.append(r)
        return "\n".join(lines)

    elif ext == "txt":
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()

    else:
        raise ValueError(f"Unsupported file type: .{ext}")

# ── CHUNKING ──────────────────────────────────────────────────────────────────

def chunk_with_metadata(text, filename):
    splitter = get_splitter()
    chunks = splitter.create_documents(
        texts=[text],
        metadatas=[{"source": filename}]
    )
    for chunk in chunks:
        chunk.page_content = f"[From: {filename}]\n" + chunk.page_content
    print(f"✅ Created {len(chunks)} chunks")
    return chunks

# ── VECTOR STORE ──────────────────────────────────────────────────────────────

def build_vector_store(chunks):
    from langchain_community.vectorstores import FAISS
    print(f"Embedding {len(chunks)} chunks...")
    vs = FAISS.from_documents(documents=chunks, embedding=get_embeddings())
    print("✅ Vector store built")
    return vs

def save_vector_store(vs, path="faiss_index"):
    vs.save_local(path)
    print("✅ Saved to disk")

def load_vector_store(path="faiss_index"):
    from langchain_community.vectorstores import FAISS
    return FAISS.load_local(
        path,
        get_embeddings(),
        allow_dangerous_deserialization=True
    )

# ── QUERY EXPANSION ───────────────────────────────────────────────────────────

def expand_query(query):
    prompt = f"""Write exactly 3 different phrasings of this question.
One per line. No numbers, no bullets, no extra text.
Question: {query}"""
    result = get_llm().invoke(prompt).content.strip()
    alternatives = [q.strip() for q in result.split("\n") if q.strip()]
    return [query] + alternatives[:3]

# ── RETRIEVE + RERANK ─────────────────────────────────────────────────────────

def retrieve_and_rerank(query, vector_store, final_k=5):
    queries = expand_query(query)
    seen, all_docs = set(), []

    for q in queries:
        for doc in vector_store.similarity_search(q, k=15):
            key = hash(doc.page_content[:80])
            if key not in seen:
                seen.add(key)
                all_docs.append(doc)

    if not all_docs:
        return []

    pairs = [[query, doc.page_content] for doc in all_docs]
    scores = get_reranker().predict(pairs)
    ranked = sorted(zip(all_docs, scores), key=lambda x: x[1], reverse=True)
    print(f"✅ {len(all_docs)} chunks → top {final_k}")
    return [doc for doc, _ in ranked[:final_k]]

# ── PROMPT ────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are a precise enterprise document assistant.
STRICT RULES:
1. Answer ONLY from the CONTEXT provided.
2. NEVER use outside knowledge.
3. If not in context say: "I could not find this in the document."
4. Quote directly when possible.
5. End every answer with — Sources: [filenames used]
"""

def build_prompt(query, docs, history=[]):
    context = "\n\n".join([
        f"[Source {i+1} — {doc.metadata.get('source','?')}]:\n{doc.page_content}"
        for i, doc in enumerate(docs)
    ])
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    messages.extend(history[-6:])
    messages.append({
        "role": "user",
        "content": f"CONTEXT:\n{context}\n\n{'='*40}\n\nQUESTION: {query}\n\nAnswer strictly from context:"
    })
    return messages

# ── VERIFY ────────────────────────────────────────────────────────────────────

def verify_answer(answer, docs, query):
    context = "\n".join([d.page_content for d in docs])[:3000]
    prompt = f"""Is this answer supported by the context?
Context: {context}
Question: {query}
Answer: {answer}
Reply exactly 2 lines:
Line 1: SUPPORTED or NOT_SUPPORTED
Line 2: One sentence reason."""
    result = get_llm().invoke(prompt).content.strip().split("\n")
    is_valid = "SUPPORTED" in result[0].upper()
    reason = result[1] if len(result) > 1 else ""
    return is_valid, reason

# ── MAIN ASK ──────────────────────────────────────────────────────────────────

def ask(query, vector_store, history=[]):
    docs = retrieve_and_rerank(query, vector_store)
    if not docs:
        return "No relevant content found in the document.", []
    messages = build_prompt(query, docs, history)
    answer = get_llm().invoke(messages).content
    is_valid, reason = verify_answer(answer, docs, query)
    if not is_valid:
        return f"Could not verify answer from document. ({reason})", []
    sources = list(set([d.metadata.get("source", "?") for d in docs]))
    return answer, sources